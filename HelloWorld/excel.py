import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):

    wb = xl.load_workbook(filename)
# the next step is case sensitive so we have to be careful
    sheet = wb['Sheet1']
# Two ways to get the same result
# cell = sheet['a1']
# cellA = sheet.cell(1, 1)
# print(cell.value)
# print(cellA.value)
# this will count the amount of rows in the spreadsheet
    print(sheet.max_row)
# now we are going to iterate through the rows
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
            min_row=2,
            max_row=sheet.max_row,
            min_col=4,
            max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')
    # we are overwriting the same file
    wb.save(filename)
